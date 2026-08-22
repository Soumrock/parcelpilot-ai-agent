from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class DataStore:
    """Loads the supplied workbook and provides scoped access to its data."""

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self.accounts = self._load_sheet("accounts")
        self.orders = self._load_sheet("orders")
        self.tickets = self._load_sheet("tickets")
        self.readme = self._load_sheet("README", dict_rows=False)

    def _load_sheet(self, sheet_name: str, dict_rows: bool = True):
        wb = load_workbook(self.workbook_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = list(rows[0])
        body = [list(row) for row in rows[1:] if any(v is not None for v in row)]
        if not dict_rows:
            return body
        records = []
        for row in body:
            record = {headers[i]: self._normalize(row[i]) if i < len(row) else None for i in range(len(headers))}
            records.append(record)
        return records

    @staticmethod
    def _normalize(value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        return value

    def authorized(self, account_id: str, user: dict) -> bool:
        return account_id in set(user.get("authorized_accounts", []))

    def get_account(self, account_id: str, user: dict) -> dict | None:
        if not self.authorized(account_id, user):
            raise PermissionError("User is not authorized to access this account.")
        return next((x for x in self.accounts if x["account_id"] == account_id), None)

    def get_order(self, order_id: str, user: dict) -> dict | None:
        order = next((x for x in self.orders if x["order_id"] == order_id), None)
        if not order:
            return None
        if not self.authorized(order["account_id"], user):
            raise PermissionError("User is not authorized to access this order.")
        return order

    def get_ticket(self, ticket_id: str, user: dict) -> dict | None:
        ticket = next((x for x in self.tickets if x["ticket_id"] == ticket_id), None)
        if not ticket:
            return None
        if not self.authorized(ticket["account_id"], user):
            raise PermissionError("User is not authorized to access this ticket.")
        return ticket

    def search(self, entity: str, query: str | None, user: dict, account_id: str | None = None) -> dict:
        source = getattr(self, entity)
        results = []
        allowed_accounts = set(user.get("authorized_accounts", []))
        for row in source:
            row_account = row.get("account_id")
            if row_account and row_account not in allowed_accounts:
                continue
            if account_id and row_account != account_id:
                continue
            if not query:
                results.append(row)
                continue
            q = query.lower()
            haystack = json.dumps(row, default=str).lower()
            if q in haystack or any(part in haystack for part in q.split() if len(part) >= 3):
                results.append(row)
        return {"count": len(results), "records": results[:50]}
